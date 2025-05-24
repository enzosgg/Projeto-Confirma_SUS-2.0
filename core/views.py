from django.http import HttpResponse
from core.models import Consulta
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font, Alignment

def gerar_estatisticas(request):
    agendada = Consulta.objects.filter(status__in=['agendada', 'confirmada']).count()
    ofertadas = Consulta.objects.filter(status__in=['aguardando', 'agendada', 'naocompareceu']).count()
    compareceu = Consulta.objects.filter(status='compareceu').count()
    perda_primaria = ofertadas - compareceu
    relacao = ((ofertadas - agendada)/ofertadas)*100 if ofertadas > 0 else 0
    naocompareceu = ((agendada - compareceu)/agendada)*100 if agendada > 0 else 0
    taxa_ocupacao = ((ofertadas - perda_primaria)/ofertadas)*100 if ofertadas > 0 else 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Estatísticas Jan-Jun"[:31]
    
    for sheet in wb.sheetnames[1:]:
        wb.remove(wb[sheet])

    ws.merge_cells('A1:J1')
    cabecalho = ws['A1']
    cabecalho.value = "Estatísticas de Consultas (Janeiro a Junho)"
    cabecalho.font = Font(bold=True, size=14)
    cabecalho.alignment = Alignment(horizontal='center', vertical='center')

    titulo = ["Ofertadas", "Agendadas", "Realizadas","Perda Primária","","Relação","Absenteísmo","","Taxa de Ocupação"]
    ws.append(titulo)
    
    letra = ["A","B","C","D","","A-B","E","","C-A"]
    ws.append(letra)
    
    ws.append([""]*10)
    
    informacoes = [ofertadas,agendada,compareceu,perda_primaria, "",f"{relacao:.1f}%",
    f"{naocompareceu:.1f}%","",f"{taxa_ocupacao:.1f}%"]
    ws.append(informacoes)
    
    for linha in ws.iter_rows(min_row=1, max_row=5):
        for celula in linha:
            if celula.row == 1:
                celula.font = Font(bold=True, size=14)
                celula.alignment = Alignment(horizontal='center', vertical='center')
            elif celula.row == 2:
                celula.font = Font(bold=True) if celula.value else Font()
                celula.alignment = Alignment(horizontal='center')
            elif celula.row == 3:
                celula.font = Font(italic=True)
                celula.alignment = Alignment(horizontal='center')
            elif celula.row == 5:
                celula.alignment = Alignment(horizontal='center')

    largura_coluna = {'A': 15,'B': 12, 'C': 18,'D': 18,'E': 5,'F': 12,'G': 15,'H': 5,'I': 16,}
    
    for letra_coluna, largura in largura_coluna.items():
        ws.column_dimensions[letra_coluna].width = largura


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        charset='utf-8'
    )
    nome_arquivo = f"estatisticas_consultas_jan-jun_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; nome_arquivo="{nome_arquivo}"'
    
    wb.save(response)
    return response